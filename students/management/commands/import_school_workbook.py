"""
Import the single-workbook Excel template (Students / Grades / Attendance sheets)
into HouseMaster's Phase 1 models.

Template contract (per school-insight-plan.md):
  Students  -> id, first_name, last_name, class, house
  Grades    -> student_id, subject, term, score, max_score
  Attendance-> student_id, date, status, notes

Usage:
  python manage.py import_school_workbook path/to/workbook.xlsx --school "Sample School"
"""
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from students.models import School, SchoolClass, YearGroup, Student
from gradebook.models import Subject, Term, Grade
from attendance.models import AttendanceRecord


def _rows(ws):
    """Yield each data row as a dict keyed by the header row (row 1)."""
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(headers, row))


class Command(BaseCommand):
    help = "Import a school's Students/Grades/Attendance Excel workbook into HouseMaster."

    def add_arguments(self, parser):
        parser.add_argument("workbook_path", type=str)
        parser.add_argument("--school", type=str, required=True, help="School name (created if it doesn't exist)")

    def handle(self, *args, **options):
        path = options["workbook_path"]
        school_name = options["school"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Workbook not found: {path}")

        school, _ = School.objects.get_or_create(name=school_name)

        students_created = self._import_students(wb, school)
        grades_created = self._import_grades(wb, school) if "Grades" in wb.sheetnames else 0
        attendance_created = self._import_attendance(wb, school) if "Attendance" in wb.sheetnames else 0

        self.stdout.write(self.style.SUCCESS(
            f"Imported for '{school.name}': "
            f"{students_created} students, {grades_created} grades, {attendance_created} attendance records."
        ))

    def _import_students(self, wb, school):
        if "Students" not in wb.sheetnames:
            raise CommandError("Workbook is missing a 'Students' sheet.")
        ws = wb["Students"]
        count = 0
        for row in _rows(ws):
            ext_id = str(row.get("id", "")).strip()
            first_name = str(row.get("first_name", "")).strip()
            last_name = str(row.get("last_name", "")).strip()
            class_name = str(row.get("class") or row.get("class/year") or "").strip()
            house = str(row.get("house") or "").strip()

            if not first_name or not last_name:
                continue  # skip malformed rows rather than hard-failing the whole import

            school_class = None
            if class_name:
                year_group, _ = YearGroup.objects.get_or_create(school=school, name=class_name)
                school_class, _ = SchoolClass.objects.get_or_create(
                    year_group=year_group, name=class_name, defaults={"house": house}
                )

            Student.objects.update_or_create(
                school=school,
                external_id=ext_id,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "school_class": school_class,
                    "house": house,
                },
            )
            count += 1
        return count

    def _import_grades(self, wb, school):
        ws = wb["Grades"]
        count = 0
        for row in _rows(ws):
            ext_id = str(row.get("student_id", "")).strip()
            subject_name = str(row.get("subject", "")).strip()
            term_name = str(row.get("term", "")).strip()
            raw_score = row.get("score")
            raw_max = row.get("max_score", 100)

            if not ext_id or not subject_name or not term_name or raw_score is None:
                continue

            try:
                student = Student.objects.get(school=school, external_id=ext_id)
            except Student.DoesNotExist:
                continue  # student not in this import batch — skip rather than fail whole run

            subject, _ = Subject.objects.get_or_create(school=school, name=subject_name)
            term, _ = Term.objects.get_or_create(school=school, name=term_name)

            try:
                score = Decimal(str(raw_score))
                max_score = Decimal(str(raw_max)) if raw_max is not None else Decimal("100")
            except InvalidOperation:
                continue

            Grade.objects.update_or_create(
                student=student, subject=subject, term=term,
                defaults={"score": score, "max_score": max_score},
            )
            count += 1
        return count

    def _import_attendance(self, wb, school):
        ws = wb["Attendance"]
        count = 0
        for row in _rows(ws):
            ext_id = str(row.get("student_id", "")).strip()
            date = row.get("date")
            status = str(row.get("status", "present")).strip().lower()
            notes = str(row.get("notes") or "").strip()

            if not ext_id or not date:
                continue

            try:
                student = Student.objects.get(school=school, external_id=ext_id)
            except Student.DoesNotExist:
                continue

            valid_statuses = {c[0] for c in AttendanceRecord.STATUS_CHOICES}
            if status not in valid_statuses:
                status = "present"

            AttendanceRecord.objects.update_or_create(
                student=student, date=date,
                defaults={"status": status, "notes": notes},
            )
            count += 1
        return count
